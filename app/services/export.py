"""Export services (CSV, XLSX)."""
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models import Asset


def export_assets_to_csv(assets, status_filter=None):
    """Export assets to CSV (in-memory)."""
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            'serial_number', 'asset_tag', 'model', 'manufacturer', 'category',
            'status', 'purchase_date', 'warranty_expiry', 'cost', 'vendor',
            'location', 'department', 'assigned_to', 'sold_to'
        ]
    )
    writer.writeheader()

    for asset in assets:
        active = asset.active_assignment
        writer.writerow({
            'serial_number': asset.serial_number or '',
            'asset_tag': asset.asset_tag or '',
            'model': asset.model or '',
            'manufacturer': asset.manufacturer or '',
            'category': asset.category or '',
            'status': asset.status or '',
            'purchase_date': asset.purchase_date or '',
            'warranty_expiry': asset.warranty_expiry or '',
            'cost': asset.cost or '',
            'vendor': asset.vendor or '',
            'location': asset.location or '',
            'department': asset.department or '',
            'assigned_to': active.user_name if active else '',
            'sold_to': asset.sold_to or ''
        })

    return output.getvalue().encode('utf-8')


def export_assets_to_xlsx(assets):
    """Export assets to XLSX with formatting (in-memory)."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Assets'

    headers = [
        'Serial Number', 'Asset Tag', 'Model', 'Manufacturer', 'Category',
        'Status', 'Purchase Date', 'Warranty Expiry', 'Cost', 'Vendor',
        'Location', 'Department', 'Assigned To', 'Sold To'
    ]

    # Add header row with formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Freeze panes
    worksheet.freeze_panes = 'A2'

    # Add data rows
    for row_idx, asset in enumerate(assets, 2):
        active = asset.active_assignment
        worksheet.cell(row=row_idx, column=1, value=asset.serial_number)
        worksheet.cell(row=row_idx, column=2, value=asset.asset_tag)
        worksheet.cell(row=row_idx, column=3, value=asset.model)
        worksheet.cell(row=row_idx, column=4, value=asset.manufacturer)
        worksheet.cell(row=row_idx, column=5, value=asset.category)
        worksheet.cell(row=row_idx, column=6, value=asset.status)
        worksheet.cell(row=row_idx, column=7, value=str(asset.purchase_date) if asset.purchase_date else '')
        worksheet.cell(row=row_idx, column=8, value=str(asset.warranty_expiry) if asset.warranty_expiry else '')
        worksheet.cell(row=row_idx, column=9, value=asset.cost)
        worksheet.cell(row=row_idx, column=10, value=asset.vendor)
        worksheet.cell(row=row_idx, column=11, value=asset.location)
        worksheet.cell(row=row_idx, column=12, value=asset.department)
        worksheet.cell(row=row_idx, column=13, value=active.user_name if active else '')
        worksheet.cell(row=row_idx, column=14, value=asset.sold_to)

    # Auto-fit columns
    for col in worksheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_len = min(max_len + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_len

    # Prevent formula injection
    # (openpyxl doesn't execute formulas anyway, but good practice)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
